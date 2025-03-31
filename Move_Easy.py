from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import filedialog
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from bs4 import BeautifulSoup
import re
import os
import subprocess
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys

# Dicionário para mapear lojas fantasias para seus IDs e CNPJs
dict_lojas = {
    "AVARE CD" : "2183783000111",
    "SOROCABA" : "2183783001606",
    "SOROCABA-MP" : "2183783001517",
    "LOJA 01 - AVARE" : "2183783000545",
    "LOJA 03 - MARILIA" : "2183783000383",
    "LOJA 04 - ARACATUBA" : "2183783000200",
    "LOJA 05 - PRUDENTE" : "2183783000464",
    "LOJA 07 - MARILIA" : "2183783000626",
    "LOJA 08 - BAURU" : "2183783000898",
    "LOJA 09 - ASSIS" : "2183783000707",
    "LOJA 10 - ANDRADINA" : "2183783000979",
    "LOJA 11 - BIRIGUI" : "2183783001193",
    "LOJA 12 - ITAPEVA" : "2183783001002",
    "LOJA 13 - BAURU SHOP" : "2183783001274",
    "LOJA 14 - OURINHOS" : "2183783001355",
    "LOJA 17 - JUNDIAI" : "14298644000112",
    "LOJA 20 - JAU" : "328944000354",
    "LOJA 21 - PINHEIROS" : "8571461000126",
    "LOJA 22 - ITAIM BIBI" : "8571461000207",
    "LOJA 25 - STA CRUZ" : "2183783001860",
    "LOJA 26 - BARAO" : "2183783001789",
    "LOJA 28 - JUNDIAI" : "2183783001940",
    "LOJA 30 - ARICANDUVA" : "2183783002165",
    "LOJA 31 - MATEO BEI" : "2183783002246",
    "LOJA 32 - TEODORO" : "2183783002327",
    "LOJA 33 - ITAIM" : "2183783002408",
    "LOJA 34 - HIGIENOPOLIS" : "2183783002599",
    "LOJA 40 - BRAGANCA I" : "72714637000150",
    "LOJA 41 - BRAGANCA II" : "72714637000401",
    "LOJA 42 - BARUERI BOUL" : "2183783003218",
    "LOJA 43 - BARUERI CAMP" : "2183783003307",
    "LOJA 44 - PERUS" : "2183783003480",
    "LOJA 45 - CRUZEIRO" : "2183783003641",
    "LOJA 46 - GUARA CENTRO" : "2183783003722",
    "LOJA 47 - GUARA SHOP" : "2183783003803",
    "LOJA 48 - LORENA" : "2183783003994",
    "LOJA 49 - PINDA" : "2183783004028",
    "LOJA 50 - TAUBATE I" : "2183783004109",
    "LOJA 51 - TAUBATE II" : "2183783004290",
    "LOJA 55 - BOTUCATU 1" : "2183783004613",
    "LOJA 56 - BOTUCATU 2" : "2183783004702",
    "LOJA 57 - JAU" : "2183783004885",
    "LOJA 58 - SOROCABA 1" : "2183783002670",
    "LOJA 60 - SAO CARLOS CENTRO" : "2183783002831",
    "LOJA 61 - SC V PRADO" : "2183783002912",
    "LOJA 62 - SC SHOPPING" : "2183783003056",
    "LOJA 63 - RIB SHOP" : "2183783003137",
    "LOJA 64 - IPIRANGA" : "2183783004966",
    "LOJA 66 - PIEDADE" : "2183783005180",
    "LOJA 67 - FRANC MORATO" : "2183783005261",
    "LOJA 68 - MARILIA III" : "2183783005423",
    "LOJA 71 - P FERREIRA" : "2183783005695",
}
dict_num_lojas = {
    "CD" : "2183783000111",
    "CD2" : "2183783001606",
    "CD1" : "2183783001517",
    "1" : "2183783000545",
    "3" : "2183783000383",
    "4" : "2183783000200",
    "5" : "2183783000464",
    "7" : "2183783000626",
    "8" : "2183783000898",
    "9" : "2183783000707",
    "10" : "2183783000979",
    "11" : "2183783001193",
    "12" : "2183783001002",
    "13" : "2183783001274",
    "14" : "2183783001355",
    "17" : "14298644000112",
    "20" : "328944000354",
    "21" : "8571461000126",
    "22" : "8571461000207",
    "25" : "2183783001860",
    "26" : "2183783001789",
    "28" : "2183783001940",
    "30" : "2183783002165",
    "31" : "2183783002246",
    "32" : "2183783002327",
    "33" : "2183783002408",
    "34" : "2183783002599",
    "40" : "72714637000150",
    "41" : "72714637000401",
    "42" : "2183783003218",
    "43" : "2183783003307",
    "44" : "2183783003480",
    "45" : "2183783003641",
    "46" : "2183783003722",
    "47" : "2183783003803",
    "48" : "2183783003994",
    "49" : "2183783004028",
    "50" : "2183783004109",
    "51" : "2183783004290",
    "55" : "2183783004613",
    "56" : "2183783004702",
    "57" : "2183783004885",
    "58" : "2183783002670",
    "60" : "2183783002831",
    "61" : "2183783002912",
    "62" : "2183783003056",
    "63" : "2183783003137",
    "64" : "2183783004966",
    "66" : "2183783005180",
    "67" : "2183783005261",
    "68" : "2183783005423",
    "71" : "2183783005695",
}
dict_grupos = {
    "CD" : "[CD] - MATRIZ",
    "20" : "[20] - LOJA 20 - JAU",
    "21" : "[21] - LOJA 21-PINHEIROS I",
    "22" : "[22] - LOJA 22-ITAIM BIBI",
    "26" : "[26] - LOJA 26-BARAO",
    "28" : "[28] - LOJA 28-JUNDIAI",
    "35" : "[35] - LOJA 35-PENHA",
    "40" : "[40] - LOJA 40-BRAGANCA I",
    "41" : "[41] - LOJA 41-BRAGANCA II",
    "42" : "[42] - LOJA 42-BARUERI BOUL",
    "52" : "[52] - LOJA 52 - COTIA II",
    "63" : "[63] - LOJA 63-RIB. SHOP",
    "67" : "[67] - LOJA 67-FRANC MORATO",
    "1" : "[1] - LISTA 0",
    "2" : "[2] - GRUPO 2",
    "3" : "[3] - GRUPO 3",
    "4" : "[4] - GRUPO 4",
    "5" : "[5] - GRUPO 5",
    "6" : "[6] - GRUPO 6"
}

# def open_file(caminho):
#     df = pd.read_excel(caminho, engine='openpyxl')
#     return df

# Função para salvar o nome de usuário e a senha em um arquivo
def save_credentials():
    username = username_entry.get()
    password = password_entry.get()
    with open('credentials.txt', 'w') as file:
        file.write(f"{username}\n{password}")

# Função para carregar o nome de usuário e a senha do arquivo
def load_credentials():
    try:
        with open('credentials.txt', 'r') as file:
            lines = file.readlines()
            username = lines[0].strip()
            password = lines[1].strip()
            username_entry.insert(0, username)
            password_entry.insert(0, password)
    except FileNotFoundError:
        # Cria o arquivo se não existir
        with open('credentials.txt', 'w') as file:
            file.write("")  # Cria um arquivo vazio

# Função para realizar o login e abrir a página principal
def login():
    save_credentials()  # Salva as credenciais
    root.destroy()  # Fecha a janela de login
    open_main_page()  # Abre a página principal

def abrir_arquivo(entrada_arquivo, root):
    arquivo_selecionado = filedialog.askopenfilename(parent=root)
    if arquivo_selecionado:
        entrada_arquivo.delete(0, tk.END)  # Limpa a entrada
        entrada_arquivo.insert(0, arquivo_selecionado)  # Insere o caminho do arquivo selecionado

def baixar_layout_alteracao():
    # Defina o nome do arquivo e o caminho para a pasta de downloads
    download_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    file_path = os.path.join(download_folder, "Layout_alteracao_preco.xlsx")
    
    # Defina os nomes das colunas conforme a imagem
    colunas = [
        "Tipo do Codigo", "Produto/Grupo", "Vl. Custo", "Vl. Revenda", "Loja/Grupo", "Data inicio" ,"Status"
    ]
    
    # Crie um dataframe vazio com as colunas desejadas
    df = pd.DataFrame(columns=colunas)
    
    df.to_excel(file_path, index=False, engine='openpyxl')

        # Ajusta o tamanho das colunas para todas terem largura 20
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    
    # Define todas as colunas com largura 20
    for col in worksheet.columns:
        col_letter = col[0].column_letter  # Pega a letra da coluna (A, B, C, etc.)
        worksheet.column_dimensions[col_letter].width = 20
    
    # Salva o arquivo após ajustar a largura das colunas
    workbook.save(file_path)
    
    # Abrir a pasta de downloads no Windows
    subprocess.Popen(f'explorer "{download_folder}"')

def baixar_layout_transferencia():
    # Defina o nome do arquivo e o caminho para a pasta de downloads
    download_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    file_path = os.path.join(download_folder, "Layout_transf_pedido.xlsx")
    
    # Defina os nomes das colunas conforme a imagem
    colunas = [
        "Loja Origem", "Loja Destino", "Qtd&Code", "Status"
    ]
    
    # Crie um dataframe vazio com as colunas desejadas
    df = pd.DataFrame(columns=colunas)
    
    df.to_excel(file_path, index=False, engine='openpyxl')

        # Ajusta o tamanho das colunas para todas terem largura 20
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    
    # Define todas as colunas com largura 20
    for col in worksheet.columns:
        col_letter = col[0].column_letter  # Pega a letra da coluna (A, B, C, etc.)
        worksheet.column_dimensions[col_letter].width = 20
    
    # Salva o arquivo após ajustar a largura das colunas
    workbook.save(file_path)
    
    # Abrir a pasta de downloads no Windows
    subprocess.Popen(f'explorer "{download_folder}"')

# Funcao para consultar o preco.
def consulta_preco(caminho, select_loja):
    import warnings
    warnings.simplefilter(action='ignore', category=FutureWarning)
    
    # Obter o CNPJ da loja selecionada
    cnpj_loja = dict_lojas.get(select_loja)
    
    def novo_caminho(caminho):
        # Extrair diretório e nome base do arquivo
        dir_name, file_name = os.path.split(caminho)
        base_name, ext = os.path.splitext(file_name)

        # Criar o novo nome de arquivo
        new_file_name = f"{base_name}_{select_loja}_preco_coletado_parcial{ext}"
        new_file_path = os.path.join(dir_name, new_file_name)

        return new_file_path

    def arquivo_final(caminho):
        # Definir o caminho do arquivo existente (parcial)
        old_file_path = novo_caminho(caminho)

        # Extrair diretório e nome base do arquivo
        dir_name, file_name = os.path.split(old_file_path)
        base_name, ext = os.path.splitext(file_name)

        # Substituir a palavra 'parcial' por 'final' no nome do arquivo
        new_file_name = file_name.replace('parcial', 'final')
        new_file_path = os.path.join(dir_name, new_file_name)
        
        # Renomear o arquivo
        os.rename(old_file_path, new_file_path)

    def abrir_arquivo_existente(caminho):
        new_file_path = novo_caminho(caminho)

        if os.path.exists(new_file_path):
            df_existente = pd.read_excel(new_file_path, dtype={'Product Code': str}, engine='openpyxl')
            return df_existente
        else:
            # Retorna um DataFrame vazio com as colunas especificadas
            colunas = ['Product Code', 'Product Description', 'Price Cust Rep', 'Price Venda',
                    'Price Promocao', 'Price Custo Cont', 'Price Fidelidade', 'Price Ecommerce']
            return pd.DataFrame(columns=colunas)  # Retorna DataFrame vazio se o arquivo não existir
    
    def enter_navegador():
        # Carregar credenciais do arquivo
        try:
            with open('credentials.txt', 'r') as file:
                lines = file.readlines()
                username = lines[0].strip()
                password = lines[1].strip()
        except:
            pass

        navegador = webdriver.Chrome()
        navegador.get("https://sumire-phd.homeip.net:8099/SistemasPHD/")
        user_name = navegador.find_element(By.ID, "form-login")
        user_password = WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located((By.ID, "form-senha"))
        )

        user_name.send_keys(username)
        user_password.send_keys(password)

        button_login = navegador.find_element(By.ID, "form-submit")
        button_login.click()

        time.sleep(2)
        navegador.find_element(By.ID, 'j_id13').click()

        time.sleep(2)
        navegador.find_element(By.ID, 'opConsultas').click()

        time.sleep(2)
        navegador.find_element(By.XPATH, "//input[starts-with(@id, 'incCentral:')]").click()

        cnpj_field = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.ID, 'incCentral:incCentralConsultas:formEscolheCnpj:selEmiCoCnpj'))
        )
        cnpj_field.click()

        select_cnpj_loja = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, f".//option[contains(@value, '{cnpj_loja}')]"))
        )
        select_cnpj_loja.click()

        return navegador

    def collect_prices(navegador, code, descricao=None):
        # Função para localizar o valor baseado no cabeçalho, pulando para a próxima linha
        def buscar_valor_proxima_linha(soup, cabecalho):
            try:
                # Encontra a <td> que contém o cabeçalho
                td_cabecalho = soup.find('td', string=cabecalho)
                tr_cabecalho = td_cabecalho.find_parent('tr')
                tds_cabecalho = tr_cabecalho.find_all('td')

                # Encontra a próxima <tr> (a linha com os valores)
                proxima_tr = tr_cabecalho.find_next('tr')

                # Captura todas as <td> da próxima linha (a linha de valores)
                tds_valores = proxima_tr.find_all('td')

                # Descobre o índice do cabeçalho desejado (ex: 'Venda')
                for i, td in enumerate(tds_cabecalho):
                    if cabecalho.strip() == td.get_text(strip=True):
                        return tds_valores[i].text.strip()
                return None
            except:
                return None

        # Função para localizar o valor do texto dentro do 'span' baseado no cabeçalho
        def localizar_valor_por_cabecalho(soup, cabecalho):
            try:
                # Procura a linha com o cabeçalho desejado
                linha = soup.find('td', string=cabecalho)             
                # Localiza o próximo 'td'
                valor_td = linha.find_next('td')
                # Busca o valor dentro do 'span'
                valor_span = valor_td.find('span')
                if valor_span:
                    return valor_span.get_text(strip=True)  # Retorna o valor visível do 'span'
                else:
                    return None
            except:
                return None
            
        if descricao is None:
            # Remover qualquer coisa que não seja número do code, se for necessário
            code = re.sub(r'\D', '', str(code).strip())  # Converte para string e mantém apenas números
            
            WebDriverWait(navegador, 20).until(
                EC.presence_of_element_located((By.ID, 'incCentral:incCentralConsultas:pnlObjQtdeValores'))
            )
            
            check_code = False
            while not check_code:
                html_content = navegador.page_source

                # Criar um objeto BeautifulSoup
                soup = BeautifulSoup(html_content, 'html.parser')
                try:
                    product_code = soup.find('td', class_='colCodigoProduto cinza1').text
                except:
                    product_code = 0
                if product_code.strip() == code:
                    check_code = True
                    break
                else:
                    pass

            # Localize os valores usando o texto dos cabeçalhos
            product_description = soup.find('td', class_='colDescricaoProduto cinza1').text
            price_cust_rep = buscar_valor_proxima_linha(soup, 'Custo Rep.')
            price_venda = buscar_valor_proxima_linha(soup, 'Venda')
            price_promocao = buscar_valor_proxima_linha(soup, 'Promoção')
            price_custo_cont = buscar_valor_proxima_linha(soup, 'Custo Cont.')
            price_fidelidade = localizar_valor_por_cabecalho(soup, 'FIDELIDADE')
            price_ecommerce = localizar_valor_por_cabecalho(soup, 'e-Commerce')

            # Função para lidar com valores None e fazer a substituição ou retornar um valor padrão
            def safe_float_conversion(value):
                if value is None:
                    return 'SEM VALOR'
                return float(value.replace(',', '.'))

            # Criação de um novo DataFrame com os dados extraídos
            new_data = {
                'Product Code': [product_code],
                'Product Description': [product_description],
                'Price Cust Rep': safe_float_conversion(price_cust_rep),
                'Price Venda': safe_float_conversion(price_venda),
                'Price Promocao': safe_float_conversion(price_promocao),
                'Price Custo Cont': safe_float_conversion(price_custo_cont),
                'Price Fidelidade': safe_float_conversion(price_fidelidade),
                'Price Ecommerce': safe_float_conversion(price_ecommerce)
            }
        else:
            new_data = {
                'Product Code': [code],
                'Product Description': [descricao],
                'Price Cust Rep': [''],
                'Price Venda': [''],
                'Price Promocao': [''],
                'Price Custo Cont': [''],
                'Price Fidelidade': [''],
                'Price Ecommerce': ['']
            }
        new_df = pd.DataFrame(new_data)

        df_frame = abrir_arquivo_existente(caminho)

        # Concatenar o novo DataFrame com o DataFrame existente
        df_frame = pd.concat([df_frame, new_df], ignore_index=True)

        new_file_path = novo_caminho(caminho)

        # Salvar o DataFrame atualizado no arquivo
        df_frame.to_excel(new_file_path, index=False)

    def selecionar_item(navegador, code):
        select_item = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//input[starts-with(@id, 'incCentral:incCentralConsultas:')]"))
        )
        select_item.click()

        try:
            # Criar um objeto BeautifulSoup
            html_content = navegador.page_source        
            soup = BeautifulSoup(html_content, 'html.parser')

            # Procurar a frase "Nenhuma linha retornada!"
            mensagem = soup.find(string=re.compile("Nenhuma linha retornada!", re.IGNORECASE))

            code_field = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.ID, 'incCentral:incCentralConsultas:formPnlModalPesquisa:txtPrdCoProdutoFiltro'))
            )
            code_field.click()
            code_field.clear()
            code_field.send_keys(code)

            consulta = WebDriverWait(navegador, 10).until(
                EC.visibility_of_element_located((By.ID, 'incCentral:incCentralConsultas:formPnlModalPesquisa:btnPsqProduto'))
            )
            consulta.click()
            time.sleep(6)
        except:
            code_field = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.ID, 'incCentral:incCentralConsultas:formPnlModalPesquisa:txtPrdCoProdutoFiltro'))
            )
            code_field.click()
            code_field.clear()
            code_field.send_keys(code)

            consulta = WebDriverWait(navegador, 10).until(
                EC.element_to_be_clickable((By.ID, 'incCentral:incCentralConsultas:formPnlModalPesquisa:btnPsqProduto'))
            )
            consulta.click()

        check_code = False
        
        # Remover qualquer coisa que não seja número do code, se for necessário
        code = re.sub(r'\D', '', str(code).strip())  # Converte para string e mantém apenas números
        while not check_code:
            tentativas =+ 1
            # Criar um objeto BeautifulSoup
            html_content = navegador.page_source        
            soup = BeautifulSoup(html_content, 'html.parser')

            # Procurar a frase "Nenhuma linha retornada!"
            mensagem = soup.find(string=re.compile("Nenhuma linha retornada!", re.IGNORECASE))

            if mensagem:
                navegador.execute_script("document.getElementById('incCentral:incCentralConsultas:pnlPsqProduto').component.hide()")
                return False
            try:
                linhas = soup.find_all('td', class_='tblLinha')
                td_linha = linhas[0] if linhas else check_code == False

                # Pegar o texto da <td> e dividir para pegar apenas o primeiro valor (número)
                texto_completo = td_linha.get_text(strip=True)

                # Extrair apenas números do code_box (caso tenha texto misturado)
                code_box = texto_completo.split(' ')[0]
                code_box = re.sub(r'\D', '', code_box.strip())  # Mantém apenas números
            except:
                code_box = 0
            # Verificar se ambos são iguais
            if code_box == code:
                check_code = True
                break
            else:
                pass

        item = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[starts-with(@id, 'incCentral:incCentralConsultas:formPnlModalPesquisa:tblPsqInfoParticipanteBody:')]"))
            )
        item.click()
        return True

    # Carregar os dados da planilha original e garantir que os códigos de produto sejam strings
    lista = pd.read_excel(caminho, dtype={0: str}, engine='openpyxl')

    # Verificar se o arquivo já existe
    df_existente = abrir_arquivo_existente(caminho)

    # Remover espaços em branco ou caracteres extras dos códigos de df_existente
    df_existente['Product Code'] = df_existente['Product Code'].astype(str).str.strip()

    # Remover espaços em branco ou caracteres extras dos códigos de lista (primeira coluna)
    codigos_novos = lista.iloc[:, 0].astype(str).str.strip()

    # Comparar a primeira coluna de 'lista' com 'Product Code' de df_existente
    codigos_existentes = df_existente['Product Code'].unique()

    # Filtrar apenas os códigos que não estão no arquivo existente
    lista = lista[~codigos_novos.isin(codigos_existentes)]

    # Inicializar o navegador e coletar dados apenas para os códigos não processados
    if not lista.empty:
        nav = enter_navegador()
        for index, row in lista.iterrows():
            code = row.iloc[0]
            codigo_localizado = selecionar_item(nav, code)
            if codigo_localizado:
                collect_prices(nav, code)
            else:
                collect_prices(nav, code, descricao='NAO LOCALIZADO')
    arquivo_final(caminho)

# Funcao para realizar a movimentacao interna
def movimentacao_interna(caminho):
    def novo_nome_csv():
        # Extrair diretório e nome base do arquivo Excel
        dir_name, file_name = os.path.split(caminho)
        base_name, ext = os.path.splitext(file_name)

        # Criar o novo nome de arquivo CSV no mesmo diretório
        csv_file_name = f"{base_name}_controle_parcial.csv"  # Altera a extensão para .csv
        csv_file_path = os.path.join(dir_name, csv_file_name)
        return csv_file_path
    
    def xml_csv():
        df = pd.read_excel(caminho, engine='openpyxl')

        # Substituir os valores na coluna da Loja Origem pelo CNPJ correspondente
        df["Loja Origem"] = df["Loja Origem"].astype(str).replace(dict_num_lojas)

        df["Loja Destino"] = df["Loja Destino"].astype(str).replace(dict_num_lojas)

        csv_file_path = novo_nome_csv()

        # Salvar o DataFrame como CSV no mesmo diretório
        df.to_csv(csv_file_path, sep=";", index=False)
        return df
    
    def leitura_planilha():
        try:
            df = pd.read_csv(novo_nome_csv(), sep=";")
        except:
            df = xml_csv()
        
        # Converter as colunas 'Loja Origem' e 'Loja Destino' para string
        df['Loja Origem'] = df['Loja Origem'].astype(str)
        df['Loja Destino'] = df['Loja Destino'].astype(str)

        # Filtrar as linhas onde o status está em branco (usando valores nulos ou strings vazias)
        df_filtrado = df[df["Status"].isna() | (df["Status"] == '')]
        
        # Separar o DataFrame em partes por 'ORIGEM' e 'DESTINO'
        grouped_dfs = df_filtrado.groupby(["Loja Origem", "Loja Destino"])

        return grouped_dfs
    
    def update_status(cnpj_origem, cnpj_destino, item, num_pedido):
        df = pd.read_csv(novo_nome_csv(), sep=";")
        
        # Converter as colunas 'Loja Origem' e 'Loja Destino' para string
        df['Loja Origem'] = df['Loja Origem'].astype(str)
        df['Loja Destino'] = df['Loja Destino'].astype(str)
        
        # Atualizar a coluna 'Status' quando as condições forem atendidas
        mask = (df['Loja Origem'] == cnpj_origem) & \
            (df['Loja Destino'] == cnpj_destino) & \
            (df['Qtd&Code'] == item)
        
        # Atualizar o status conforme a máscara
        df.loc[mask, 'Status'] = f"PD {num_pedido}"

        # Salvar o DataFrame de volta ao CSV (substituindo o arquivo original)
        df.to_csv(novo_nome_csv(), sep=";", index=False)


    def arquivo_final():
        # Definir o caminho do arquivo existente (parcial)
        old_file_path = novo_nome_csv()

        # Extrair diretório e nome base do arquivo
        dir_name, file_name = os.path.split(old_file_path)
        base_name, ext = os.path.splitext(file_name)

        # Substituir a palavra 'parcial' por 'final' no nome do arquivo
        new_file_name = file_name.replace('parcial', 'final')
        new_file_path = os.path.join(dir_name, new_file_name)
        
        # Renomear o arquivo
        os.rename(old_file_path, new_file_path)

    def enter_navegador():
        # Carregar credenciais do arquivo
        try:
            with open('credentials.txt', 'r') as file:
                lines = file.readlines()
                username = lines[0].strip()
                password = lines[1].strip()
        except:
            pass
        
        # Entrar no navegador
        navegador = webdriver.Chrome()
        navegador.get("https://sumire-phd.homeip.net:8099/eVendas/home.faces")
        user_name = navegador.find_element(By.ID, "form-login")
        user_password = WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located((By.ID, "form-senha"))
        )

        # Credenciais
        user_name.send_keys(username)
        user_password.send_keys(password)

        button_login = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.ID, "form-submit"))
        )
        button_login.click()

        return navegador


    def enter_mov_int(navegador):
        navegador.find_element(By.ID, 'opMovInterna').click()

        select_field = WebDriverWait(navegador, 5).until(
            EC.presence_of_element_located((By.ID, 'incCentral:formConteudo:selPadraoLancamento'))
        )
        select = Select(select_field)
        select.select_by_visible_text('[TRANSFERENCIA]')

        iniciar = navegador.find_element(By.ID, 'incCentral:formConteudo:btnIniciar')
        iniciar.click()

        return navegador

    def action_mov_int(navegador,cnpj_origem, cnpj_destino, lista):
        select_cnpj = WebDriverWait(navegador, 5).until(
            EC.element_to_be_clickable((By.XPATH, f".//option[contains(@value, '{cnpj_origem}')]"))
        )
        select_cnpj.click()

        current_value = 0
        
        for row in lista:
            quantite_field = navegador.find_element(By.ID, 'incCentral:formConteudo:txtProduto')
            quantite_field.click()
            quantite_field.clear()
            quantite_field.send_keys(row)
           
            navegador.find_element(By.ID, 'incCentral:formConteudo:btnAdicionar').click()
            
            quantite = int(row.split('&')[0])
            current_value += quantite

            text_value = 0

            while current_value != text_value:
                time.sleep(2)
                html_content = navegador.page_source

                # Criar um objeto BeautifulSoup
                soup_content = BeautifulSoup(html_content, 'html.parser')

                # Usar BeautifulSoup para encontrar o elemento desejado
                item_span = soup_content.find('span', string='ITENS:')
                
                text_value = int(item_span.next_sibling.split(' ')[-1])
                
        navegador.find_element(By.ID, 'incCentral:formConteudo:btnPesqCliente').click()

        time.sleep(2)
        cnpj = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.ID, 'incCentral:formPnlModalPesquisaParticipante:txtPtcCoCnpjFiltro'))
        )
        cnpj.click()
        cnpj.send_keys(cnpj_destino)

        navegador.find_element(By.ID, 'incCentral:formPnlModalPesquisaParticipante:btnPsqPtcControlado').click()

        select_id_destino = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[starts-with(@id, 'incCentral:formPnlModalPesquisaParticipante:tblPsqInfoPtcControladoBody:0:')]"))
        )
        select_id_destino.click()

        cond_pag = WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located((By.ID, 'incCentral:formConteudo:selCondicaoPagto'))
        )
        select_pagamento = Select(cond_pag)
        select_pagamento.select_by_visible_text('[100] - TRANSFERENCIA')

        navegador.find_element(By.ID, 'incCentral:formConteudo:btnFinalizar').click()

        time.sleep(3)
        # Captura o HTML do elemento
        html_element = navegador.page_source

        soup_element = BeautifulSoup(html_element, 'html.parser')

        item_class = soup_element.find('li', class_='okMessageGrande')
        
        value_element = item_class.text.split(' ')[2]

        for row in lista:
            item = row
            update_status(cnpj_origem, cnpj_destino, item, value_element)

        # Selecionar Vendas
        navegador.find_element(By.ID, 'opVendas').click()

        # Selecionar Pedidos
        navegador.find_element(By.XPATH, "//*[contains(@id, 'opPedidoVenda')]").click()

        # Habilitar o campo CNPJ
        habilitar_cnpj = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.ID, 'incCentral:incCentralVenda:formConteudo:formEmitente:selFiltroEmiCoCnpj'))
        )
        habilitar_cnpj.click()

        # Selecionar o CNPJ
        select_cnpj2 = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, f".//option[contains(@value, '{cnpj_origem}')]"))
        )
        select_cnpj2.click()

        # Pesquisar
        navegador.find_element(By.ID, 'incCentral:incCentralVenda:formConteudo:btnPedPesquisar').click()

        # Localizar posicao do numero do pedido
        num_order = WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//td[text()= '{value_element}']"))
        )
        position = num_order.location
        y_num_order = position['y']

        # Localizar todos os elementos "Pastel" que seguem o padrão no XPath
        elementos_pastel = navegador.find_elements(By.XPATH, "//*[starts-with(@id, 'incCentral:incCentralVenda:formConteudo:tblPrdBodyPesquisa:')]")

        # Iterar sobre todos os elementos localizados
        for elemento in elementos_pastel:
            # Pegar a localização do elemento "Pastel"
            posicao_pastel = elemento.location
            y_pastel = posicao_pastel['y']

            # Comparar a coordenada y do pedido com o elemento pastel
            if y_pastel == y_num_order + 2:
                # Se as posições y forem iguais, clique no elemento
                elemento.click()
                break  # Para o loop após encontrar e clicar no elemento correto

        # Clicar em Liberar Faturamento
        faturamento = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.ID, 'incCentral:incCentralVenda:formConteudo:btnPedLiberar'))
        )
        faturamento.click()

        time.sleep(2)
        # Esperar até que o alerta apareça e interagir com ele
        alerta = Alert(navegador)

        # Aceitar o alerta clicando no botão "OK"
        alerta.accept()

        time.sleep(3)

        # Selecionar Faturamento
        navegador.find_element(By.XPATH, "//*[contains(@id, 'opFaturamento')]").click()

        # Habilitar o campo CNPJ
        habilitar_cnpj2 = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.ID, 'incCentral:incCentralVenda:formConteudo:formEmitente:selFiltroEmiCoCnpj'))
        )
        habilitar_cnpj2.click()

        # Selecionar o CNPJ
        select_cnpj3 = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, f".//option[contains(@value, '{cnpj_origem}')]"))
        )
        select_cnpj3.click()

        # Pesquisar
        navegador.find_element(By.ID, 'incCentral:incCentralVenda:formConteudo:btnPdfPesquisar').click()

        # Localizar posicao do numero do pedido
        num_order = WebDriverWait(navegador, 20).until(
            EC.presence_of_element_located((By.XPATH, f"//td[text()= '{value_element}']"))
        )
        position = num_order.location
        y_num_order = position['y']

        # Localizar todos os elementos "Operacao" que seguem o padrão no XPath
        elementos_operacao = navegador.find_elements(By.XPATH, "//*[contains(@id, 'btnSelOperacaoPed')]")

        # Iterar sobre todos os elementos localizados
        for elemento in elementos_operacao:
            # Pegar a localização do elemento "Operacao"
            posicao_operacao = elemento.location
            y_operacao = posicao_operacao['y']

            # Comparar a coordenada y do pedido com o elemento operacao
            if y_operacao == y_num_order + 2:
                # Se as posições y forem iguais, clique no elemento
                elemento.click()
                break  # Para o loop após encontrar e clicar no elemento correto

        time.sleep(2)
        # Esperar até que o alerta apareça e interagir com ele
        alerta = Alert(navegador)

        # Aceitar o alerta clicando no botão "OK"
        alerta.accept()

        time.sleep(1)
        # Selecionar operacao
        navegador.find_element(By.ID, 'incCentral:incCentralVenda:frmmodalOperacao:rgnSelOperacaomodalOperacao').click()

        # Aguarde o select aparecer
        field_ope = WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[starts-with(@id, 'incCentral:incCentralVenda:frmmodalOperacao:selOpeCoOperacao')]"))
        )

        # Re-encontrar o elemento antes de usar Select
        select_ope = Select(field_ope)
        select_ope.select_by_visible_text('[10] - TRANSFERENCIA DE MERCADORIA')

        navegador.find_element(By.XPATH, "//input[@value='Selecionar Operação']").click()

        time.sleep(5)

    def processo_inclusao_pedidos():
        navegador = enter_navegador()
        grouped_dfs = leitura_planilha()
        for group_name, group_df in grouped_dfs:
            origem, destino = group_name  # 'group_name' retorna uma tupla com (Loja Origem, Loja Destino)
            lista = group_df['Qtd&Code'].tolist()  # Converte a coluna 'Qtd&Code' para uma lista
            enter_mov_int(navegador)
            action_mov_int(navegador, origem, destino, lista)
        arquivo_final()

    processo_inclusao_pedidos()

# Funcao para alterar o preco
def alteracao_preco(caminho):
    import warnings
    warnings.simplefilter(action='ignore', category=FutureWarning)
    
    def loading(navegador):
        time.sleep(1)
        # Espera até que o status de carregamento mude para "display: none"
        WebDriverWait(navegador, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span[id='_viewRoot:status.start'][style='display: none']"))
        )

    def enter_navegador():
        # Carregar credenciais do arquivo
        with open('credentials.txt', 'r') as file:
            lines = file.readlines()
            username = lines[0].strip()
            password = lines[1].strip()

        navegador = webdriver.Chrome()
        navegador.get("https://sumire-phd.homeip.net:8099/eVendas/home.faces")
        wait = WebDriverWait(navegador,10)
        user_name = navegador.find_element(By.ID, "form-login")
        user_password = wait.until(EC.presence_of_element_located((By.ID, "form-senha")))

        user_name.send_keys(username)
        user_password.send_keys(password)

        button_login = navegador.find_element(By.ID, "form-submit")
        button_login.click()

        field_cadastro = wait.until(EC.element_to_be_selected((By.ID, 'op2')))
        field_cadastro.click()

        field_precos = wait.until(EC.element_to_be_selected((By.ID, 'op78')))
        field_precos.click()

        field_manutencao = wait.until(EC.element_to_be_selected((By.ID, 'op58')))
        field_manutencao.click()
        return navegador
    
    def box_message_nenhuma(navegador):
        try:
            # Espera até que o elemento com a classe 'divNenhumaLinha' esteja presente
            element_nenhuma_linha = navegador.find_element(By.XPATH, "//*[contains(@class, 'divNenhumaLinha')]")
            # Aqui você pode interagir com o elemento ou retornar o texto, por exemplo
            texto = element_nenhuma_linha.text
            return texto
        except:
            return None
    
    def box_message_td(navegador, tipo):
        if tipo == "grupo":
            try:
                # Tenta capturar o texto no primeiro XPath id="incCentral:formPnlModalPesquisaGpr:txtGrpCoGrupoFiltro"
                element_tds = navegador.find_elements(By.XPATH, "//*[@id='incCentral:formPnlModalPesquisaGpr:txtGrpCoGrupoFiltro']//td[@class='tblLinha']")
                if element_tds:
                    descricao = element_tds[0].text
                    texto = str(descricao.split(' - ')[0])
                    return texto
            except:
                return None
        else:
            try:
                # Se não encontrar o primeiro elemento, tenta capturar o segundo XPath
                element_tds = navegador.find_elements(By.XPATH, "//*[@id='incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaPrd:tblPsqInfoParticipanteBody']//td[@class='tblLinha']")
                if element_tds:
                    descricao = element_tds[0].text
                    texto = str(descricao.split(' - ')[0])
                    return texto
            except:
                return None
    
    def get_value(navegador, tipo):
        if tipo == "grupo":
            try:
                # Encontra o elemento pelo ID
                elemento_input = navegador.find_element(By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaGpr:txtGrpCoGrupoFiltro")
                if elemento_input:
                    # Obtém o valor do atributo 'value'
                    valor = elemento_input.get_attribute('value')
                    return str(valor)
            except:
                return None
        else:
            try:
                elemento_input = navegador.find_element(By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaPrd:txtPrdCoProdutoFiltro")
                if elemento_input:
                    # Obtém o valor do atributo 'value'
                    valor = elemento_input.get_attribute('value')
                    return str(valor)
            except:
                return None

    def capturar_codigo_span(navegador):
        try:
            # Localiza o elemento span pelo texto exato
            span_element = navegador.find_element(By.XPATH, "//span[contains(text(), 'Produto pertence a um grupo de preços. Não permite alterar!')]")

            # Obtém o texto do elemento
            span_text = span_element.text

            # Usa uma expressão regular para capturar o número entre parênteses
            match = re.search(r'\((\d+)\)', span_text)
            
            if match:
                codigo = match.group(1)  # Extrai o número capturado
                return codigo
            else:
                return None
        except:
            return None
    
    def seleciona_loja(navegador, num_loja):
        time.sleep(1)
        try:
            loja = dict_grupos.get(num_loja)
            element_loja = WebDriverWait(navegador,10).until(EC.visibility_of_element_located((By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formConteudo:selEmiCoCnpj")))
            select_loja = Select(element_loja)
            select_loja.select_by_visible_text(loja)
            return navegador, True
        except:
            return navegador, False
    
    def inclui_data_inicio(navegador, data):
        field_data = WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.XPATH, "//*[contains(@id, 'txtPvdDtIniValidade')]")))
        # Usar JavaScript para limpar o campo
        navegador.execute_script("arguments[0].value = '';", field_data)
        
        # Usar JavaScript para definir o valor no campo
        navegador.execute_script("arguments[0].value = arguments[1];", field_data, data)
        
        # Disparar o evento 'change' para garantir que o valor seja registrado
        navegador.execute_script("arguments[0].dispatchEvent(new Event('change'));", field_data)
        return navegador

    def selecionar_grupo_preco(navegador, code, tipo):
        loading(navegador)
        code = str(code)
        
        # Clica no botão para selecionar o grupo de preço
        button_grupo_preco = WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.XPATH, "//input[@value='Sel. o Grupo Preço']")))
        button_grupo_preco.click()
        
        time.sleep(2)

        while True:
            if get_value(navegador, tipo) == "":
                break
            if box_message_nenhuma(navegador) is not None or box_message_td(navegador, tipo) is not None:
                break
            time.sleep(1)

        value = get_value(navegador, tipo)
        first_message_td = box_message_td(navegador, tipo)
        first_message_nenhum = box_message_nenhuma(navegador)

        if value == code:
            if first_message_nenhum != None:
                navegador.execute_script("document.getElementById('incCentral:incCentralDiversos:incCentralDiversos:pnlPsqGrupoPreco').component.hide()")
                return navegador, f"ERRO - {first_message_nenhum}"
            else:             
                pastel_all_codes = WebDriverWait(navegador,10).until(EC.presence_of_element_located((By.XPATH, "//*[starts-with(@id, 'incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaGpr:tblPsqInfoGrupoPrecoBody:')]")))
                pastel_codes = pastel_all_codes.find_elements(By.XPATH, "//*[starts-with(@id, 'incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaGpr:tblPsqInfoGrupoPrecoBody:')]")
                # Executa o clique no segundo elemento usando JavaScript
                navegador.execute_script("arguments[0].click();", pastel_codes[1])
                return navegador, first_message_td

        field_code_grupe = WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaGpr:txtGrpCoGrupoFiltro")))
        field_code_grupe.click()
        field_code_grupe.clear()
        field_code_grupe.send_keys(code)

        button_pesquisar = WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaGpr:btnPsqGrupoPreco")))
        button_pesquisar.click()

        time.sleep(3)

        while True:
            if box_message_td(navegador, tipo) is not None:
                if code == box_message_td(navegador, tipo):
                    break
            if box_message_nenhuma(navegador) is not None:
                break
            time.sleep(1)
        
        mensagem_td = box_message_td(navegador, tipo)
        mensagem_nenhuma = box_message_nenhuma(navegador)

        if mensagem_nenhuma is not None:
            navegador.execute_script("document.getElementById('incCentral:incCentralDiversos:incCentralDiversos:pnlPsqGrupoPreco').component.hide()")
            return navegador, f"ERRO - {mensagem_nenhuma}"
        else:
            if code == mensagem_td:
                pastel_all_codes = WebDriverWait(navegador,10).until(EC.presence_of_element_located((By.XPATH, "//*[starts-with(@id, 'incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaGpr:tblPsqInfoGrupoPrecoBody:')]")))
                pastel_codes = pastel_all_codes.find_elements(By.XPATH, "//*[starts-with(@id, 'incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaGpr:tblPsqInfoGrupoPrecoBody:')]")
                # Executa o clique no segundo elemento usando JavaScript
                navegador.execute_script("arguments[0].click();", pastel_codes[1])
                return navegador, mensagem_td
            else:
                navegador.execute_script("document.getElementById('incCentral:incCentralDiversos:incCentralDiversos:pnlPsqGrupoPreco').component.hide()")
                return navegador, "ERRO"

    def selecionar_produto(navegador, code, tipo):
        loading(navegador)
        code = str(code)
        
        # Clica no botão para selecionar o produto
        button_produto = WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.XPATH, "//input[@value='Selecionar o Produto']")))
        button_produto.click()

        time.sleep(2)

        while True:
            if get_value(navegador, tipo) == "":
                break
            if box_message_nenhuma(navegador) is not None or box_message_td(navegador, tipo) is not None:
                break
            time.sleep(1)

        first_message_td = box_message_td(navegador, tipo)
        first_message_nenhum = box_message_nenhuma(navegador)
        value = get_value(navegador, tipo)

        if value == code:
            if first_message_nenhum != None:
                navegador.execute_script("document.getElementById('incCentral:incCentralDiversos:incCentralDiversos:pnlPsqProduto').component.hide()")
                return navegador, f"ERRO - {first_message_nenhum}"
            else:
                # Espera até que o elemento seja clicável incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaPrd:tblPsqInfoParticipanteBody:0:j_id278
                pastel_all_codes= WebDriverWait(navegador,10).until(EC.presence_of_element_located((By.XPATH, "//*[starts-with(@id, 'incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaPrd:tblPsqInfoParticipanteBody:')]")))
                # Aqui estamos buscando os mesmos elementos que o 'pastel_all_codes' se refere
                pastel_codes = pastel_all_codes.find_elements(By.XPATH, "//*[starts-with(@id, 'incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaPrd:tblPsqInfoParticipanteBody:')]")
                # Executa o clique no segundo elemento usando JavaScript
                navegador.execute_script("arguments[0].click();", pastel_codes[1])
                time.sleep(2)
                span = capturar_codigo_span(navegador)
                if span != None:
                    return navegador, f"ERRO - Grupo de preco {span}"
                return navegador, first_message_td

        field_code_grupe = WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaPrd:txtPrdCoProdutoFiltro")))
        field_code_grupe.click()
        field_code_grupe.clear()
        field_code_grupe.send_keys(code)

        button_pesquisar = WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaPrd:btnPsqProduto")))
        button_pesquisar.click()

        time.sleep(3)

        while True:
            if box_message_td(navegador, tipo) is not None:
                if code == box_message_td(navegador, tipo):
                    break
            if box_message_nenhuma(navegador) is not None:
                break
            time.sleep(1)
    
        mensagem_td = box_message_td(navegador, tipo)
        mensagem_nenhuma = box_message_nenhuma(navegador)

        if mensagem_nenhuma is not None:
            navegador.execute_script("document.getElementById('incCentral:incCentralDiversos:incCentralDiversos:pnlPsqProduto').component.hide()")
            return navegador, f"ERRO - {mensagem_nenhuma}"
        else:
            if mensagem_td == code:
                    # Espera até que o elemento seja clicável incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaPrd:tblPsqInfoParticipanteBody:0:j_id278
                    pastel_all_codes = WebDriverWait(navegador,10).until(EC.presence_of_element_located((By.XPATH, "//*[starts-with(@id, 'incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaPrd:tblPsqInfoParticipanteBody:')]")))
                    # Aqui estamos buscando os mesmos elementos que o 'pastel_all_codes' se refere
                    pastel_codes = pastel_all_codes.find_elements(By.XPATH, "//*[starts-with(@id, 'incCentral:incCentralDiversos:incCentralDiversos:formPnlModalPesquisaPrd:tblPsqInfoParticipanteBody:')]")
                    # Executa o clique no segundo elemento usando JavaScript
                    navegador.execute_script("arguments[0].click();", pastel_codes[1])
                    time.sleep(2)
                    span = capturar_codigo_span(navegador)
                    if span != None:
                        return navegador, f"ERRO - {span}"
                    return navegador, mensagem_td
            else:
                navegador.execute_script("document.getElementById('incCentral:incCentralDiversos:incCentralDiversos:pnlPsqProduto').component.hide()")
                return navegador, "ERRO"
            
    def validacao_dados(navegador):
        WebDriverWait(navegador,10).until(EC.visibility_of_element_located((By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formConteudo:pnlPrecoVendaForm")))
        start_time_loja = time.time()
        check_loja = ''
        while True:
            # Localiza o elemento <select> pelo ID 
            select_element = navegador.find_element(By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formConteudo:selEmiCoCnpj")

            # Localiza o <option> que está selecionado
            selected_option = select_element.find_element(By.XPATH, "./option[@selected='selected']")

            # Obtém o texto do <option> selecionado
            option_text = selected_option.text
            # Usa uma expressão regular para capturar o número entre colchetes
            match = re.search(r'\[(.*?)\]', option_text)

            if match:
                check_loja = match.group(1)  # Extrai o valor capturado
                break
            
            if time.time() - start_time_loja > 5:
                break
        
        start_time_code = time.time()
        check_code = ''
        while True:
            # Localiza o elemento span usando o ID
            span_element = navegador.find_element(By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formConteudo:pnlNoPrdFiltro")

            # Obtém o texto do elemento
            span_text = span_element.text

            # Usa uma expressão regular para capturar o número dentro dos colchetes
            match = re.search(r'\[(\d+)\]', span_text)
            
            if match:
                check_code = str(match.group(1))  # Extrai o número capturado
                break
            if time.time() - start_time_code > 5:
                break
        
        return check_loja, check_code

    def atualiza_preco(navegador, vl_custo, vl_revenda):
        try:
            field_custo = navegador.find_element(By.ID, 'incCentral:incCentralDiversos:incCentralDiversos:formConteudo:txtPvdVlCustoReposicao')
        except:
            field_custo = navegador.find_element(By.ID, 'incCentral:incCentralDiversos:incCentralDiversos:formConteudo:txtGpvVlCustoReposicao')
        field_custo.click()
        field_custo.clear()
        field_custo.send_keys(vl_custo)

        try:
            field_revenda = navegador.find_element(By.ID, 'incCentral:incCentralDiversos:incCentralDiversos:formConteudo:txtPvdVlVendaRevenda')
        except:
            field_revenda = navegador.find_element(By.ID, 'incCentral:incCentralDiversos:incCentralDiversos:formConteudo:txtGpvVlVendaRevenda')
        field_revenda.click()
        field_revenda.clear()
        field_revenda.send_keys(vl_revenda)

        button_salvar = navegador.find_element(By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formConteudo:btngpvPrecoAddEdt")
        button_salvar.click()

        start_time_save = time.time()
        while True:
            # Aguardar até que o elemento seja encontrado ou o tempo limite seja atingido
            ul_element = WebDriverWait(navegador, 20).until(EC.presence_of_element_located((By.ID, "incCentral:incCentralDiversos:incCentralDiversos:formConteudo:msgEndGlobal")))

            # Localiza o <li> com a classe 'okMessage' dentro do <ul>
            li_element = ul_element.find_element(By.CLASS_NAME, "okMessage")

            # Obtém o texto do <li>
            mensagem_ok = li_element.text

            if mensagem_ok == 'Salvo com sucesso!':
                return navegador, True
            if time.time() - start_time_save > 10:
                return navegador, False

    def novo_nome_csv():
        # Extrair diretório e nome base do arquivo Excel
        dir_name, file_name = os.path.split(caminho)
        base_name, ext = os.path.splitext(file_name)

        # Criar o novo nome de arquivo CSV no mesmo diretório
        csv_file_name = f"{base_name}_alteracao_preco_parcial.csv"  # Altera a extensão para .csv
        csv_file_path = os.path.join(dir_name, csv_file_name)
        return csv_file_path
    
    def xml_csv():
        # Carregar a planilha
        df = pd.read_excel(caminho, engine='openpyxl')
        
        # Normalizar o case da coluna 'Tipo do Codigo' para aceitar qualquer valor independente de maiúsculas/minúsculas
        df['Tipo do Codigo'] = df['Tipo do Codigo'].str.lower()  # Converte para minúsculas para padronizar

        try:
            # Substituir vírgula por ponto em 'Vl. Custo' e 'Vl. Revenda' se necessário
            df['Vl. Custo'] = df['Vl. Custo'].astype(str).str.replace(',', '.').astype(float)
        except:
            pass
        try:
            df['Vl. Revenda'] = df['Vl. Revenda'].astype(str).str.replace(',', '.').astype(float)
        except:
            pass
        try:
            # Converter a coluna 'Data inicio' para datetime e depois para o formato desejado
            df['Data inicio'] = pd.to_datetime(df['Data inicio'], errors='coerce')  # Converter para datetime
            df['Data inicio'] = df['Data inicio'].dt.strftime('%d/%m/%Y')  # Formatar como dia/mês/ano
        except:
            pass

        csv_file_path = novo_nome_csv()

        # Salvar o DataFrame como CSV no mesmo diretório
        df.to_csv(csv_file_path, sep=";", index=False)
        return df

    def arquivo_final():
        # Definir o caminho do arquivo existente (parcial)
        old_file_path = novo_nome_csv()

        # Extrair diretório e nome base do arquivo
        dir_name, file_name = os.path.split(old_file_path)
        base_name, ext = os.path.splitext(file_name)

        # Substituir a palavra 'parcial' por 'final' no nome do arquivo
        new_file_name = file_name.replace('parcial', 'final')
        new_file_path = os.path.join(dir_name, new_file_name)
        
        # Renomear o arquivo
        os.rename(old_file_path, new_file_path)

    def analisar_linha(df):
        navegador = enter_navegador()
        # Iterar sobre as linhas da planilha
        for idx, row in df.iterrows():
            # Verificar o status da linha
            data = row['Data inicio']
            status = str(row['Status'])
            codigo = row['Produto/Grupo']
            lojas_total = row['Loja/Grupo']
            lojas = str(lojas_total).split(',')
            tipo = row['Tipo do Codigo']
            vl_custo = row['Vl. Custo']
            vl_revenda = row['Vl. Revenda']

            if status.startswith("OK") or status.startswith("ERRO"):
                # Se o status for OK ou ERRO, pular para a próxima linha
                continue

            elif status.startswith("PARCIAL"):
                # Se o status for PARCIAL, obter as lojas já analisadas
                lojas_analisadas = status.split('-')[-1].split(',')
                # Encontrar a próxima loja/grupo a ser analisada
                lojas_pendentes = [loja for loja in lojas if loja not in lojas_analisadas]
                analisadas = status
                if len(lojas_pendentes) == 0:
                    df.at[idx, 'Status'] = "OK"
                    df.to_csv(novo_nome_csv(), sep=";" ,index=False)
                    continue
            else:
                # Se não for PARCIAL, todas as lojas/grupos estão pendentes
                lojas_pendentes = lojas
                analisadas = "PARCIAL-"

            inclui_data_inicio(navegador, data)

            # Verificar se ainda há lojas/grupos pendentes
            if lojas_pendentes:
                for loja in lojas_pendentes:
                    navegador, bool_loja = seleciona_loja(navegador, loja)
                    if bool_loja is False:
                        analisadas_parcial = analisadas.split("-")[0] + f" Loja {loja} nao localizada."
                        if analisadas.split("-")[-1]:
                            analisadas = analisadas_parcial + "-" + analisadas.split("-")[-1]
                        else:
                            analisadas = analisadas_parcial + "-"
                        df.at[idx, 'Status'] = analisadas
                        df.to_csv(novo_nome_csv(), sep=";" ,index=False)
                        continue
                    else:
                        if tipo == 'produto':
                            navegador, mensagem = selecionar_produto(navegador, codigo, tipo)
                        else:
                            navegador, mensagem = selecionar_grupo_preco(navegador, codigo, tipo)
                        # Atualizar o status com a mensagem retornada
                        if mensagem.startswith("ERRO"):
                            df.at[idx, 'Status'] = mensagem
                            df.to_csv(novo_nome_csv(), sep=";" ,index=False)
                            break
                        else:
                            check_loja, check_code = validacao_dados(navegador)
                            if check_loja == loja and check_code == str(codigo):
                                navegador, bool_status = atualiza_preco(navegador, vl_custo, vl_revenda)
                                if bool_status is False:
                                    continue
                                else:
                                    if analisadas.split("-")[-1]:
                                        analisadas = analisadas + "," + loja
                                    else:
                                        analisadas = analisadas + loja
                                    df.at[idx, 'Status'] = analisadas
                                    df.to_csv(novo_nome_csv(), sep=";" ,index=False)
                            else:
                                continue

                new_status = analisadas.split("-")[-1]
                if lojas_total == new_status:
                    df.at[idx, 'Status'] = "OK"
                    df.to_csv(novo_nome_csv(), sep=";" ,index=False)

    def analisar_planilha():
        try:
            df = pd.read_csv(novo_nome_csv(), sep=";")
        except:
            df = xml_csv()
    
        # Verificar se há algum status "PARCIAL" ou em branco
        if df['Status'].isnull().any() or any(df['Status'].str.startswith("PARCIAL")):
            
            # Rodar a função para continuar o processamento
            analisar_linha(df)
        else:
            # Se todos os status forem "OK" ou "ERRO", rodar arquivo_final()
            arquivo_final()
            

    analisar_planilha()

# Função para abrir a página principal
def open_main_page():
    # Criação da Página Principal
    main_page = Tk()
    main_page.title('Página Principal')
    main_page.configure(background='DeepPink2')
    main_page.geometry("500x200")
    main_page.resizable(True, True)
    main_page.minsize(width=500, height=200)

    frame1 = Frame(main_page)
    frame1.place(relx=0.015, rely=0.03, relwidth=0.97, relheight=0.94)

    buttom_consulta_preco = Button(frame1, text="Consulta Preço", bd=3, command=lambda: page_search_price(main_page, dict_lojas))
    buttom_consulta_preco.place(relx=0.2, rely=0.3, relwidth=0.2, relheight=0.3)
    buttom_mov_interna = Button(frame1, text="Mov. Interna", bd=3, command=lambda: page_mov_int(main_page, dict_lojas))
    buttom_mov_interna.place(relx=0.4, rely=0.3, relwidth=0.2, relheight=0.3)
    buttom_novo = Button(frame1, text="Alteraçao Preço", bd=3, command=lambda: page_alt_price(main_page))
    buttom_novo.place(relx=0.6, rely=0.3, relwidth=0.2, relheight=0.3)

    main_page.mainloop()

# Criação da Página de Movimentação Interna
def page_mov_int(main_page, dict_lojas):
    root1 = Toplevel(main_page)
    root1.title("Movimentação Interna")
    root1.configure(background='DeepPink2')
    root1.geometry("500x200")
    root1.resizable(True, True)
    root1.minsize(width=500, height=200)

    # Pegar a posição da main_page
    x_main_page = main_page.winfo_x()
    y_main_page = main_page.winfo_y()

    # Definir root1 para ser aberta no mesmo local de main_page
    root1.geometry(f"+{x_main_page}+{y_main_page}")

    # Tornar root1 uma janela "filha" de main_page
    root1.transient(main_page)
    root1.lift()
    root1.grab_set()

    # Frame
    frame3 = Frame(root1)
    frame3.place(relx=0.015, rely=0.03, relwidth=0.97, relheight=0.94)

    # Inclusão do arquivo de entrada
    folder_name = Label(frame3, text="Arquivo")
    folder_name.place(relx=0.0, rely=0.1, relwidth=0.3, relheight=0.1)
    folder = Entry(frame3)
    folder.place(relx=0.30, rely=0.1, relwidth=0.5, relheight=0.1)
    buttom_search_folder = tk.Button(frame3, text='...', command=lambda: abrir_arquivo(folder, root1))
    buttom_search_folder.place(relx=0.8, rely=0.1, relwidth=0.1, relheight=0.1)

    # Inclusao do botao Baixar Layout
    buttom_layout = Button(frame3, text="Layout", bd=3, command=baixar_layout_transferencia)
    buttom_layout.place(relx=0.5, rely=0.75, relwidth=0.2, relheight=0.15)

    # Inclusão do botão Iniciar
    buttom_start = Button(frame3, text="Iniciar", bd=3, command=lambda: movimentacao_interna(folder.get()))
    buttom_start.place(relx=0.7, rely=0.75, relwidth=0.2, relheight=0.15)

# Criação da Página de Consulta Preco
def page_search_price(main_page, dict_lojas):
    root1 = Toplevel(main_page)
    root1.title("Consulta de Precos")
    root1.configure(background='DeepPink2')
    root1.geometry("500x200")
    root1.resizable(True, True)
    root1.minsize(width=500, height=200)

    # Pegar a posição da main_page
    x_main_page = main_page.winfo_x()
    y_main_page = main_page.winfo_y()

    # Definir root1 para ser aberta no mesmo local de main_page
    root1.geometry(f"+{x_main_page}+{y_main_page}")

    # Tornar root1 uma janela "filha" de main_page
    root1.transient(main_page)
    root1.lift()
    root1.grab_set()

    # Frame
    frame3 = Frame(root1)
    frame3.place(relx=0.015, rely=0.03, relwidth=0.97, relheight=0.94)

    # Inclusão do arquivo de entrada
    folder_name = Label(frame3, text="Arquivo")
    folder_name.place(relx=0.0, rely=0.1, relwidth=0.3, relheight=0.1)
    folder = Entry(frame3)
    folder.place(relx=0.30, rely=0.1, relwidth=0.6, relheight=0.1)
    buttom_search_folder = tk.Button(frame3, text='...', command=lambda: abrir_arquivo(folder, root1))
    buttom_search_folder.place(relx=0.8, rely=0.1, relwidth=0.1, relheight=0.1)

    # Inclusão Combo Box Selecionar Loja
    select_loja_label = Label(frame3, text="Selecionar Loja")
    select_loja_label.place(relx=0.0, rely=0.35, relwidth=0.25, relheight=0.1)
    select_loja = ttk.Combobox(frame3, values=list(dict_lojas.keys()))
    select_loja.place(relx=0.3, rely=0.35, relwidth=0.6, relheight=0.1)

    # Inclusão do botão Iniciar
    buttom_start = Button(frame3, text="Iniciar", bd=3, command=lambda: consulta_preco(folder.get(), select_loja.get()))
    buttom_start.place(relx=0.7, rely=0.75, relwidth=0.2, relheight=0.15)

# Criacao da pagina Alteracao de Preco
def page_alt_price(main_page):
    root1 = Toplevel(main_page)
    root1.title("Alteração de Preços")
    root1.configure(background='DeepPink2')
    root1.geometry("500x200")
    root1.resizable(True, True)
    root1.minsize(width=500, height=200)

    # Pegar a posição da main_page
    x_main_page = main_page.winfo_x()
    y_main_page = main_page.winfo_y()

    # Definir root1 para ser aberta no mesmo local de main_page
    root1.geometry(f"+{x_main_page}+{y_main_page}")

    # Tornar root1 uma janela "filha" de main_page
    root1.transient(main_page)
    root1.lift()
    root1.grab_set()

    # Frame
    frame3 = Frame(root1)
    frame3.place(relx=0.015, rely=0.03, relwidth=0.97, relheight=0.94)

    # Inclusão do arquivo de entrada
    folder_name = Label(frame3, text="Arquivo")
    folder_name.place(relx=0.0, rely=0.1, relwidth=0.3, relheight=0.1)
    folder = Entry(frame3)
    folder.place(relx=0.30, rely=0.1, relwidth=0.6, relheight=0.1)
    buttom_search_folder = tk.Button(frame3, text='...', command=lambda: abrir_arquivo(folder, root1))
    buttom_search_folder.place(relx=0.8, rely=0.1, relwidth=0.1, relheight=0.1)

    # Inclusao do botao Baixar Layout
    buttom_layout = Button(frame3, text="Layout", bd=3, command=baixar_layout_alteracao)
    buttom_layout.place(relx=0.5, rely=0.75, relwidth=0.2, relheight=0.15)

    # Inclusão do botão Iniciar
    buttom_start = Button(frame3, text="Iniciar", bd=3, command=lambda: alteracao_preco(folder.get()))
    buttom_start.place(relx=0.7, rely=0.75, relwidth=0.2, relheight=0.15)


# Janela de login
root = Tk()
root.title('Login')
root.configure(background='DeepPink2')
root.geometry("500x200")
root.resizable(True, True)
root.minsize(width=500, height=200)

frame1 = Frame(root)
frame1.place(relx=0.015, rely=0.03, relwidth=0.97, relheight=0.94)

# Nome de usuário
Label(frame1, text="Nome de Usuário").place(relx=0.2, rely=0.2)
username_entry = Entry(frame1)
username_entry.place(relx=0.5, rely=0.2)

# Senha
Label(frame1, text="Senha").place(relx=0.2, rely=0.4)
password_entry = Entry(frame1, show="*")
password_entry.place(relx=0.5, rely=0.4)

# Botão para logar
login_button = Button(frame1, text="Logar", command=login)
login_button.place(relx=0.5, rely=0.6, relwidth=0.2)

# Carregar credenciais ao iniciar o programa
load_credentials()

root.mainloop()